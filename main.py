import requests
import hashlib
import openpyxl
from os import environ
from dotenv import load_dotenv

#
# prepping API credentials
#
load_dotenv()

sid = environ.get('API_SID')
token = environ.get('API_TOKEN')

# generating md5 hash ... original hex val 'token' -> encoding to byte '.encode()' -> hash byte val '.md5()' -> encode hashed byte val to hex val '.hexdigest()'
md5_token = (hashlib.md5(token.encode())).hexdigest()


#
# creating and opening an excel file, then accessing the active sheet
#
workbook = openpyxl.Workbook()
sheet = workbook.active


#
# setting/retrieving data needed for API usage
#
current_page = 1
items_per_page = 50
dict_keys = ['name', 'cssi_id', 'custom_price', 'map_price', 'drop_ship_price', 'retail_price', 'retail_map_price']
payload = {'page' : current_page, 'per_page' : items_per_page}

# connecting to API using custom auth header instead of requests HTTPBasicAuth method
response = requests.get(url='https://api.chattanoogashooting.com/rest/v4/items', headers={"Authorization":f'Basic {sid}:{md5_token}'}, params=payload)

# making a single request to get the amount of pages in the database
data = response.json()
page_count = data["pagination"]["page_count"]


#
# retrieving item data from API and populating spreadsheet
#

# note: rows/columns start at 1, not 0

# adding titles to columns
row_val = 2
col_val = 2
for key in dict_keys:
    current_cell = sheet.cell(row = row_val, column = col_val)
    current_cell.value = key
    col_val += 1

# adding API data to cells
row_val = 3
col_val = 2
while (current_page <= page_count):

    # retrieve items for current page
    payload = {'page' : current_page, 'per_page' : items_per_page}
    response = requests.get(url='https://api.chattanoogashooting.com/rest/v4/items', headers={"Authorization":f'Basic {sid}:{md5_token}'}, params=payload)
    data = response.json()

    # loop through items in page, add to spreadsheet (if drop shippable)
    for item in data['items']:
        if (item['drop_ship_flag'] == 1):
            for key in dict_keys:
                current_cell = sheet.cell(row = row_val, column = col_val)
                current_cell.value = item[key]
                col_val += 1
            row_val += 1
            col_val = 2
    
    # increment page number
    current_page += 1

# write out the excel file
workbook.save(filename="test.xlsx")
